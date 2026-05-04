#!/usr/bin/env python3
"""
Create a sample Excel roster file for testing the oral performance evaluation system.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Sample student data
students = [
    ("Ahmed", "Hassan", "S001", "ahmed.hassan@example.com", "B1"),
    ("Fatima", "Ali", "S002", "fatima.ali@example.com", "B2"),
    ("Mohammed", "Omar", "S003", "mohammed.omar@example.com", "A2"),
    ("Layla", "Ibrahim", "S004", "layla.ibrahim@example.com", "C1"),
    ("Karim", "Noor", "S005", "karim.noor@example.com", "B2"),
    ("Noor", "Salim", "S006", "noor.salim@example.com", "A1"),
    ("Sara", "Khalil", "S007", "sara.khalil@example.com", "B1"),
    ("Hamza", "Rashid", "S008", "hamza.rashid@example.com", "C2"),
]

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Students"

# Style definitions
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal="center", vertical="center")

# Add headers
headers = ["firstName", "lastName", "studentId", "email", "cefr"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_align

# Add student data
for row_idx, (first_name, last_name, student_id, email, cefr) in enumerate(students, 2):
    ws.cell(row=row_idx, column=1, value=first_name).border = border
    ws.cell(row=row_idx, column=2, value=last_name).border = border
    ws.cell(row=row_idx, column=3, value=student_id).border = border
    ws.cell(row=row_idx, column=4, value=email).border = border
    ws.cell(row=row_idx, column=5, value=cefr).border = border

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 10

# Save the file
output_path = Path(__file__).parent / "sample-roster.xlsx"
wb.save(output_path)
print(f"✅ Sample roster file created: {output_path}")
print(f"📊 Rows: {len(students)} students + 1 header row")
